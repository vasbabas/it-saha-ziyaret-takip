"""
reports.py
----------
IT Saha Ziyaret Takip Uygulaması - Rapor ve Dışa Aktarım Modülü
Excel (.xlsx) ve PDF formatında rapor üretimi.
"""

import io
import pandas as pd
from datetime import datetime

# ---------------------------------------------------------------------------
# Excel Dışa Aktarım
# ---------------------------------------------------------------------------

def export_to_excel(df: pd.DataFrame, title: str = "Ziyaret Raporu") -> bytes:
    """
    Pandas DataFrame'ini biçimlendirilmiş Excel dosyasına (.xlsx) dönüştürür.
    
    Özellikler:
    - Başlık satırı (koyu mavi arka plan, beyaz yazı)
    - Alternatif satır renklendirmesi (zebra banding)
    - Otomatik sütun genişliği ayarı
    - Tüm kenarlıklar

    Args:
        df    : Dışa aktarılacak veri çerçevesi
        title : Rapor başlığı (sayfa adı olarak da kullanılır)
    
    Returns:
        Excel dosyasının byte içeriği (indirme için)
    """
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Sayfa adı max 31 karakter

    # --- Renkler ---
    HEADER_BG   = "1E3A5F"    # Koyu lacivert
    HEADER_FG   = "FFFFFF"    # Beyaz
    ROW_ODD     = "EBF2FA"    # Açık mavi-gri
    ROW_EVEN    = "FFFFFF"    # Beyaz
    BORDER_CLR  = "B8C6D6"    # Gri kenarlık

    # Kenarlık stili
    thin = Side(style="thin", color=BORDER_CLR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # --- Rapor Başlık Satırı (1. Satır) ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"🔷 {title}  —  Oluşturulma: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    title_cell.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="0D2137")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # --- Sütun Başlıkları (2. Satır) ---
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font      = Font(name="Calibri", bold=True, size=11, color=HEADER_FG)
        cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
    ws.row_dimensions[2].height = 22

    # --- Veri Satırları ---
    for row_idx, row_data in enumerate(df.itertuples(index=False), start=3):
        bg_color = ROW_ODD if (row_idx % 2 == 1) else ROW_EVEN
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill      = PatternFill("solid", fgColor=bg_color)
            cell.font      = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border    = border
        ws.row_dimensions[row_idx].height = 18

    # --- Otomatik Sütun Genişliği ---
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(col_name)),
            *(len(str(v)) for v in df.iloc[:, col_idx - 1].fillna("").astype(str))
        ) if len(df) > 0 else len(str(col_name))
        # Genişliği sınırlandır: min 12, max 50 karakter
        ws.column_dimensions[col_letter].width = min(50, max(12, max_len + 3))

    # --- Pencereyi dondur (başlık satırları sabit kalsın) ---
    ws.freeze_panes = "A3"

    # Belleğe yaz ve byte döndür
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ---------------------------------------------------------------------------
# PDF Dışa Aktarım
# ---------------------------------------------------------------------------

def export_to_pdf(df: pd.DataFrame, title: str = "Ziyaret Raporu") -> bytes:
    """
    Pandas DataFrame'ini biçimlendirilmiş PDF dosyasına dönüştürür.
    
    Özellikler:
    - Şirket başlığı ve rapor tarihi
    - Zebra banding'li tablo
    - Türkçe karakter desteği (ReportLab Helvetica latin1)
    - Otomatik sayfa sonu
    - A4 yatay (landscape) veya dikey

    Args:
        df    : Dışa aktarılacak veri çerçevesi
        title : Rapor başlığı

    Returns:
        PDF dosyasının byte içeriği
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    output = io.BytesIO()

    # Sütun sayısına göre yatay/dikey sayfa seç
    page_size = landscape(A4) if len(df.columns) > 5 else A4
    doc = SimpleDocTemplate(
        output,
        pagesize=page_size,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()

    # Özel stil tanımları
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Title"],
        fontSize=16,
        textColor=colors.HexColor("#1E3A5F"),
        spaceAfter=6,
        alignment=TA_CENTER,
    )
    subtitle_style = ParagraphStyle(
        "Subtitle",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.HexColor("#666666"),
        spaceAfter=16,
        alignment=TA_CENTER,
    )
    footer_style = ParagraphStyle(
        "Footer",
        parent=styles["Normal"],
        fontSize=8,
        textColor=colors.HexColor("#999999"),
        spaceBefore=10,
        alignment=TA_CENTER,
    )

    story = []

    # Başlık bölümü
    story.append(Paragraph(_safe_text(title), title_style))
    story.append(Paragraph(
        f"Oluşturulma Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}  |  Toplam Kayıt: {len(df)}",
        subtitle_style,
    ))
    story.append(Spacer(1, 0.3 * cm))

    # Tablo verisi hazırla
    # Sütun başlıkları
    header_row = [_safe_text(col) for col in df.columns]
    data = [header_row]

    for _, row in df.iterrows():
        data.append([_safe_text(str(v) if pd.notna(v) else "") for v in row])

    # Sayfa genişliğine göre sütun genişliklerini orantıla
    available_width = (landscape(A4)[0] if len(df.columns) > 5 else A4[0]) - 3 * cm
    col_width = available_width / len(df.columns)
    col_widths = [col_width] * len(df.columns)

    # Notlar sütunu varsa ona daha fazla alan ver
    for i, col in enumerate(df.columns):
        if "not" in col.lower() or "açıklama" in col.lower() or "konu" in col.lower():
            col_widths[i] = col_width * 2.0
    # Oranı yeniden normalize et
    scale = available_width / sum(col_widths)
    col_widths = [w * scale for w in col_widths]

    tbl = Table(data, colWidths=col_widths, repeatRows=1)

    # Tablo stili
    tbl_style = TableStyle([
        # Başlık satırı
        ("BACKGROUND",   (0, 0), (-1, 0),  colors.HexColor("#1E3A5F")),
        ("TEXTCOLOR",    (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",     (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",     (0, 0), (-1, 0),  9),
        ("ALIGN",        (0, 0), (-1, 0),  "CENTER"),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUND",(0, 1), (-1, -1), [colors.HexColor("#EBF2FA"), colors.white]),
        # Veri satırları
        ("FONTNAME",     (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",     (0, 1), (-1, -1), 8),
        ("ALIGN",        (0, 1), (-1, -1), "LEFT"),
        # Kenarlıklar
        ("GRID",         (0, 0), (-1, -1), 0.5, colors.HexColor("#B8C6D6")),
        ("LINEBELOW",    (0, 0), (-1, 0),  1.5, colors.HexColor("#1E3A5F")),
        # Satır yüksekliği ve padding
        ("TOPPADDING",   (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
        ("LEFTPADDING",  (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("WORDWRAP",     (0, 0), (-1, -1), True),
    ])

    # Tek çift satır renklendirme (zebra)
    for row_idx in range(1, len(data)):
        bg = colors.HexColor("#EBF2FA") if row_idx % 2 == 0 else colors.white
        tbl_style.add("BACKGROUND", (0, row_idx), (-1, row_idx), bg)

    tbl.setStyle(tbl_style)
    story.append(tbl)
    story.append(Spacer(1, 0.5 * cm))
    story.append(Paragraph(
        "IT Saha Ziyaret ve Takip Otomasyonu — Gizlidir",
        footer_style,
    ))

    doc.build(story)
    return output.getvalue()


def fmt_date(s: str) -> str:
    try:
        return datetime.strptime(s, "%Y-%m-%d").strftime("%d.%m.%Y")
    except:
        return s or "—"


def _safe_text(text: str) -> str:
    """
    Türkçe karakterleri PDF için Latin-1 uyumlu hale getirir.
    ReportLab'ın standart fontları UTF-8'i tam desteklemez.
    """
    replacements = {
        "ı": "i", "İ": "I",
        "ğ": "g", "Ğ": "G",
        "ü": "u", "Ü": "U",
        "ş": "s", "Ş": "S",
        "ö": "o", "Ö": "O",
        "ç": "c", "Ç": "C",
    }
    for tr_char, latin_char in replacements.items():
        text = text.replace(tr_char, latin_char)
    return text


def build_dataframe(visits: list[dict]) -> pd.DataFrame:
    """
    Ham ziyaret kayıtları listesini, rapora uygun Türkçe sütun başlıklı
    DataFrame'e dönüştürür.
    
    Args:
        visits: database.get_visits() çıktısı

    Returns:
        Biçimlendirilmiş pandas DataFrame
    """
    if not visits:
        return pd.DataFrame()

    df = pd.DataFrame(visits)

    # Gereksiz dahili sütunları çıkar
    drop_cols = ["id", "created_at"]
    df.drop(columns=[c for c in drop_cols if c in df.columns], inplace=True)

    # Tarih biçimlendirme
    if "visit_date" in df.columns:
        df["visit_date"] = pd.to_datetime(df["visit_date"]).dt.strftime("%d.%m.%Y")

    # Süre biçimlendirme
    if "duration" in df.columns:
        df["duration"] = df["duration"].apply(lambda x: f"{x:.1f} saat" if x else "—")

    # Sütun adlarını Türkçeye çevir
    col_rename = {
        "visit_date":  "Tarih",
        "company":     "Firma",
        "contact":     "İletişim Kişisi",
        "subject":     "Konu",
        "work_notes":  "Yapılan İşlemler",
        "duration":    "Süre",
        "technician":  "Kategori",
        "status":      "Durum",
    }
    df.rename(columns=col_rename, inplace=True)

    return df


def export_executive_pdf(visits: list[dict], stats: dict, title: str = "IT Faaliyet ve Performans Sunumu") -> bytes:
    """
    Kisisel calismalari yoneticiye sunmak uzere kapakli,
    ozet istatistik kartli ve şık teknik detay tablolu
    ust duzey bir PDF sunum raporu uretir.
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, KeepTogether
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    output = io.BytesIO()

    # Dikey A4 plani, ferah kenar bosluklari
    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        leftMargin=2.0 * cm,
        rightMargin=2.0 * cm,
        topMargin=2.2 * cm,
        bottomMargin=2.2 * cm,
    )

    styles = getSampleStyleSheet()

    # Stil Tanimlari
    style_title = ParagraphStyle(
        "ExecTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=20,
        textColor=colors.HexColor("#1A3A5F"),
        alignment=TA_CENTER,
        spaceAfter=15,
    )
    style_subtitle = ParagraphStyle(
        "ExecSub",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=10,
        textColor=colors.HexColor("#556677"),
        alignment=TA_CENTER,
        spaceAfter=30,
    )
    style_h1 = ParagraphStyle(
        "ExecH1",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=13,
        textColor=colors.HexColor("#1A3A5F"),
        spaceBefore=14,
        spaceAfter=8,
        borderPadding=4,
    )
    style_body = ParagraphStyle(
        "ExecBody",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        textColor=colors.HexColor("#334455"),
        leading=13,
    )
    style_table_header = ParagraphStyle(
        "ExecTableHeader",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=9,
        textColor=colors.white,
        alignment=TA_CENTER,
    )
    style_table_cell = ParagraphStyle(
        "ExecTableCell",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=11,
    )
    style_kpi_val = ParagraphStyle(
        "ExecKpiVal",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=16,
        textColor=colors.HexColor("#1E88E5"),
        alignment=TA_CENTER,
    )
    style_kpi_lbl = ParagraphStyle(
        "ExecKpiLbl",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
        textColor=colors.HexColor("#556677"),
        alignment=TA_CENTER,
    )

    story = []

    # -----------------------------------------------------------------------
    # 1. SAYFA: RAPOR KAPAGI & BANNERS
    # -----------------------------------------------------------------------
    story.append(Spacer(1, 1.5 * cm))
    
    # Degrade benzeri şık bir üst logo/banner alani
    banner_data = [[Paragraph("🔷 PERSONAL IT ACTIVITY DIARY", ParagraphStyle("BannerTxt", fontName="Helvetica-Bold", fontSize=11, textColor=colors.white, alignment=TA_CENTER))]]
    banner_tbl = Table(banner_data, colWidths=[17 * cm])
    banner_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#1A3A5F")),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
    ]))
    story.append(banner_tbl)
    story.append(Spacer(1, 2.5 * cm))

    # Ana Baslik
    story.append(Paragraph(_safe_text(title), style_title))
    story.append(Paragraph(
        f"Hazirlayan: IT Destek & Sistem Uzmani  |  Tarih: {datetime.now().strftime('%d.%m.%Y')}",
        style_subtitle
    ))
    story.append(Spacer(1, 1.5 * cm))

    # -----------------------------------------------------------------------
    # 2. ANA METRIKLER / KPI KARTLARI (Kapak Sayfasinda Alt Kisim)
    # -----------------------------------------------------------------------
    story.append(Paragraph(_safe_text("📊 DÖNEMSEL PERFORMANS VE ZAMAN ÖZETİ"), style_h1))
    
    # 4 Sutunlu KPI Tablosu (Kart Gorunumu)
    kpi_data = [
        [
            Paragraph(str(stats.get("total_visits", len(visits))), style_kpi_val),
            Paragraph(f"{stats.get('total_hours', 0.0):.1f} sa", style_kpi_val),
            Paragraph(str(stats.get("unique_companies", 0)), style_kpi_val),
            Paragraph(f"{stats.get('month_visits', 0)}", style_kpi_val)
        ],
        [
            Paragraph(_safe_text("Toplam Faaliyet"), style_kpi_lbl),
            Paragraph(_safe_text("Toplam Mesai"), style_kpi_lbl),
            Paragraph(_safe_text("Firma / Bölüm"), style_kpi_lbl),
            Paragraph(_safe_text("Bu Ayki Kayit"), style_kpi_lbl)
        ]
    ]
    
    kpi_tbl = Table(kpi_data, colWidths=[4.25 * cm] * 4)
    kpi_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F5F8FA")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 1, colors.HexColor("#E2E8F0")),
        ("TOPPADDING", (0, 0), (-1, -1), 12),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
    ]))
    story.append(kpi_tbl)
    
    # Alt Not
    story.append(Spacer(1, 1.0 * cm))
    story.append(Paragraph(
        _safe_text("Bu rapor; IT departmani veya kurum BT altyapisinda yapilan teknik calismalari belgelemek, "
                   "harcanan zaman analizini yapmak ve ust yonetime seffaflik saglamak amaciyla uretilmistir."),
        style_body
    ))
    
    story.append(PageBreak()) # Detaylara gecmeden once yeni sayfaya at

    # -----------------------------------------------------------------------
    # 3. DETAYLI FAALIYET TABLOSU (Sayfa 2+)
    # -----------------------------------------------------------------------
    story.append(Paragraph(_safe_text("📋 GERÇEKLEŞTİRİLEN IT FAALİYET DETAYLARI"), style_h1))
    story.append(Spacer(1, 0.2 * cm))

    # Basliklar
    headers = [
        Paragraph(_safe_text("Tarih"), style_table_header),
        Paragraph(_safe_text("Firma / Bölüm"), style_table_header),
        Paragraph(_safe_text("Konu / Islem"), style_table_header),
        Paragraph(_safe_text("Teknik Detaylar ve Yapilan Islemler"), style_table_header),
        Paragraph(_safe_text("Kategori"), style_table_header),
        Paragraph(_safe_text("Sure"), style_table_header)
    ]
    
    table_rows = [headers]

    for v in visits:
        table_rows.append([
            Paragraph(_safe_text(fmt_date(v["visit_date"])), style_table_cell),
            Paragraph(_safe_text(v["company"]), style_table_cell),
            Paragraph(_safe_text(v.get("subject") or "Genel Calisma"), style_table_cell),
            Paragraph(_safe_text(v.get("work_notes") or ""), style_table_cell),
            Paragraph(_safe_text(v.get("technician") or "Diger IT Isleri"), style_table_cell),
            Paragraph(_safe_text(f"{v.get('duration', 0.0):.1f} sa"), style_table_cell)
        ])

    # Genislikleri ayarla
    col_widths = [1.8 * cm, 2.8 * cm, 2.8 * cm, 6.0 * cm, 2.4 * cm, 1.2 * cm]
    
    tbl = Table(table_rows, colWidths=col_widths, repeatRows=1)
    
    # Şık, kurumsal tablo stili
    tbl_style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A3A5F")),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ])
    
    # Alternatif satirlari renklendir
    for idx in range(1, len(table_rows)):
        bg = colors.HexColor("#F8FAFC") if idx % 2 == 0 else colors.white
        tbl_style.add("BACKGROUND", (0, idx), (-1, idx), bg)
        
    tbl.setStyle(tbl_style)
    story.append(tbl)

    # Rapor Sonu Imzasi / Kapatis
    story.append(Spacer(1, 1.2 * cm))
    
    footer_data = [
        [
            Paragraph("", style_body),
            Paragraph(_safe_text("<b>Raporu Sunan:</b><br/>IT Sistem ve Destek Uzmani<br/>Imza / Onay"), ParagraphStyle("RightBody", parent=style_body, alignment=TA_RIGHT))
        ]
    ]
    footer_tbl = Table(footer_data, colWidths=[10 * cm, 7 * cm])
    footer_tbl.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
    ]))
    story.append(KeepTogether(footer_tbl))

    doc.build(story)
    return output.getvalue()


def send_email_report(smtp_server: str, smtp_port: str, smtp_user: str, smtp_password: str, from_email: str, to_email: str, use_ssl: bool, subject: str, body_text: str, pdf_data: bytes = None, pdf_filename: str = "rapor.pdf"):
    """
    SMTP sunucusu araciligiyla e-posta raporu gonderir. Opsiyonel olarak PDF eki ekler.
    """
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.application import MIMEApplication

    try:
        msg = MIMEMultipart()
        msg['From'] = from_email
        msg['To'] = to_email
        msg['Subject'] = subject
        
        # Turkce karakter destegi icin utf-8 ayari ile metni ekle
        msg.attach(MIMEText(body_text, 'plain', 'utf-8'))
        
        if pdf_data is not None:
            part = MIMEApplication(pdf_data, Name=pdf_filename)
            part['Content-Disposition'] = f'attachment; filename="{pdf_filename}"'
            msg.attach(part)
            
        if use_ssl:
            server = smtplib.SMTP_SSL(smtp_server, int(smtp_port), timeout=15)
        else:
            server = smtplib.SMTP(smtp_server, int(smtp_port), timeout=15)
            server.starttls()
            
        server.login(smtp_user, smtp_password)
        # Birden fazla alici varsa virgulle ayrilmis e-postalari listeye cevir
        recipients = [email.strip() for email in to_email.split(",") if email.strip()]
        server.sendmail(from_email, recipients, msg.as_string())
        server.quit()
        return True, "E-posta raporu basariyla gonderildi!"
    except Exception as e:
        return False, f"E-posta gonderilirken hata olustu: {str(e)}"

